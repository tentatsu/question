import os, sys, glob
import random
sys.path.append(os.path.join(os.path.dirname(__file__), 'site-packages'))
import openpyxl as px

EXCEL_FILE = 'excels/'
def make():
    r = random.randrange(250,900)

    for i in list(range(1000)):
        wb = px.Workbook()
        ws = wb.worksheets[0]
        if r == i:
            ws.cell(row=1, column=1, value="当たり！")
        else:
            ws.cell(row=1, column=1, value="はずれ！")
        wb.save('{}{}.xlsx'.format(EXCEL_FILE,str(i).zfill(4)))

if __name__ == '__main__':
    make()

